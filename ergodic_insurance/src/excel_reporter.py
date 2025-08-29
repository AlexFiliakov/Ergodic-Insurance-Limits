"""Excel report generation for financial statements and analysis.

This module provides comprehensive Excel report generation functionality,
creating professional financial statements, diagnostic reports, and
Monte Carlo aggregations with advanced formatting and validation.

Example:
    Generate Excel report from simulation::

        from ergodic_insurance.src.excel_reporter import ExcelReporter, ExcelReportConfig
        from ergodic_insurance.src.manufacturer import WidgetManufacturer

        # Configure report
        config = ExcelReportConfig(
            output_path=Path("./reports"),
            include_balance_sheet=True,
            include_income_statement=True,
            include_cash_flow=True
        )

        # Generate report
        reporter = ExcelReporter(config)
        output_file = reporter.generate_trajectory_report(
            manufacturer,
            "financial_statements.xlsx"
        )
"""

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, List, Optional, Tuple, Union
import warnings

import numpy as np
import pandas as pd

# Try to import Excel libraries with fallback
try:
    import xlsxwriter
    from xlsxwriter.utility import xl_col_to_name, xl_range

    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False
    warnings.warn("XlsxWriter not available. Some formatting features may be limited.")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn("openpyxl not available. Using pandas default Excel writer.")

from .financial_statements import (
    FinancialStatementConfig,
    FinancialStatementGenerator,
    MonteCarloStatementAggregator,
)

if TYPE_CHECKING:
    from .manufacturer import WidgetManufacturer


@dataclass
class ExcelReportConfig:
    """Configuration for Excel report generation.

    Attributes:
        output_path: Directory for output files
        include_balance_sheet: Whether to include balance sheet
        include_income_statement: Whether to include income statement
        include_cash_flow: Whether to include cash flow statement
        include_reconciliation: Whether to include reconciliation sheet
        include_metrics_dashboard: Whether to include metrics dashboard
        include_pivot_data: Whether to include pivot-ready data sheet
        formatting: Custom formatting options
        engine: Excel engine to use ('xlsxwriter', 'openpyxl', 'auto')
        currency_format: Currency format string
        decimal_places: Number of decimal places for numbers
        date_format: Date format string
    """

    output_path: Path = field(default_factory=lambda: Path("./reports"))
    include_balance_sheet: bool = True
    include_income_statement: bool = True
    include_cash_flow: bool = True
    include_reconciliation: bool = True
    include_metrics_dashboard: bool = True
    include_pivot_data: bool = True
    formatting: Optional[Dict[str, Any]] = None
    engine: str = "auto"
    currency_format: str = "$#,##0"
    decimal_places: int = 0
    date_format: str = "yyyy-mm-dd"


class ExcelReporter:
    """Main Excel report generation engine.

    This class handles the creation of comprehensive Excel reports
    from simulation data, including financial statements, metrics
    dashboards, and reconciliation reports.

    Attributes:
        config: Report configuration
        workbook: Excel workbook object
        formats: Dictionary of Excel format objects
        engine: Selected Excel engine
    """

    def __init__(self, config: Optional[ExcelReportConfig] = None):
        """Initialize Excel reporter.

        Args:
            config: Report configuration
        """
        self.config = config or ExcelReportConfig()
        self.workbook: Optional[Any] = None
        self.formats: Dict[str, Any] = {}

        # Select Excel engine
        self._select_engine()

        # Ensure output directory exists
        self.config.output_path.mkdir(parents=True, exist_ok=True)

    def _select_engine(self) -> None:
        """Select the Excel engine based on availability and configuration."""
        if self.config.engine == "xlsxwriter" and XLSXWRITER_AVAILABLE:
            self.engine = "xlsxwriter"
        elif self.config.engine == "openpyxl" and OPENPYXL_AVAILABLE:
            self.engine = "openpyxl"
        elif self.config.engine == "auto":
            if XLSXWRITER_AVAILABLE:
                self.engine = "xlsxwriter"
            elif OPENPYXL_AVAILABLE:
                self.engine = "openpyxl"
            else:
                self.engine = "pandas"
                warnings.warn("No Excel library available. Using pandas default writer.")
        else:
            self.engine = "pandas"

    def generate_trajectory_report(
        self, manufacturer: "WidgetManufacturer", output_file: str, title: Optional[str] = None
    ) -> Path:
        """Generate Excel report for a single simulation trajectory.

        Creates a comprehensive Excel workbook with financial statements,
        metrics, and reconciliation for a single simulation run.

        Args:
            manufacturer: WidgetManufacturer with simulation data
            output_file: Name of output Excel file
            title: Optional report title

        Returns:
            Path to generated Excel file
        """
        output_path = self.config.output_path / output_file

        # Create statement generator
        stmt_config = FinancialStatementConfig(
            currency_symbol=self.config.currency_format[0] if self.config.currency_format else "$",
            decimal_places=self.config.decimal_places,
        )
        generator = FinancialStatementGenerator(manufacturer=manufacturer, config=stmt_config)

        if self.engine == "xlsxwriter":
            self._generate_with_xlsxwriter(generator, output_path, title)
        elif self.engine == "openpyxl":
            self._generate_with_openpyxl(generator, output_path, title)
        else:
            self._generate_with_pandas(generator, output_path)

        return output_path

    def _generate_with_xlsxwriter(
        self, generator: FinancialStatementGenerator, output_path: Path, title: Optional[str] = None
    ) -> None:
        """Generate report using XlsxWriter engine.

        Args:
            generator: Financial statement generator
            output_path: Output file path
            title: Optional report title
        """
        import xlsxwriter

        # Create workbook
        self.workbook = xlsxwriter.Workbook(str(output_path))
        self._setup_xlsxwriter_formats()

        # Add cover sheet
        if title:
            self._add_cover_sheet_xlsxwriter(title)

        # Add financial statements
        if self.config.include_balance_sheet:
            self._write_balance_sheets_xlsxwriter(generator)

        if self.config.include_income_statement:
            self._write_income_statements_xlsxwriter(generator)

        if self.config.include_cash_flow:
            self._write_cash_flows_xlsxwriter(generator)

        if self.config.include_reconciliation:
            self._write_reconciliation_xlsxwriter(generator)

        if self.config.include_metrics_dashboard:
            self._write_metrics_dashboard_xlsxwriter(generator)

        if self.config.include_pivot_data:
            self._write_pivot_data_xlsxwriter(generator)

        # Close workbook
        self.workbook.close()

    def _setup_xlsxwriter_formats(self) -> None:
        """Setup XlsxWriter formatting styles."""
        assert self.workbook is not None, "Workbook must be initialized"
        self.formats = {
            "currency": self.workbook.add_format(
                {"num_format": self.config.currency_format, "align": "right"}
            ),
            "currency_bold": self.workbook.add_format(
                {"num_format": self.config.currency_format, "align": "right", "bold": True}
            ),
            "percent": self.workbook.add_format({"num_format": "0.0%", "align": "right"}),
            "number": self.workbook.add_format({"num_format": "#,##0", "align": "right"}),
            "header": self.workbook.add_format(
                {
                    "bold": True,
                    "bg_color": "#D3D3D3",
                    "border": 1,
                    "align": "center",
                    "valign": "vcenter",
                }
            ),
            "subheader": self.workbook.add_format(
                {"bold": True, "bg_color": "#F0F0F0", "border": 1}
            ),
            "total": self.workbook.add_format(
                {"bold": True, "top": 2, "bottom": 6, "num_format": self.config.currency_format}
            ),
            "subtotal": self.workbook.add_format(
                {"bold": True, "top": 1, "num_format": self.config.currency_format}
            ),
            "title": self.workbook.add_format({"bold": True, "font_size": 14, "align": "center"}),
            "date": self.workbook.add_format(
                {"num_format": self.config.date_format, "align": "center"}
            ),
            "text": self.workbook.add_format({"align": "left"}),
            "text_indent": self.workbook.add_format({"align": "left", "indent": 2}),
            "section_header": self.workbook.add_format({"bold": True, "font_size": 12}),
            "good": self.workbook.add_format({"bg_color": "#C6EFCE", "font_color": "#006100"}),
            "bad": self.workbook.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"}),
            "neutral": self.workbook.add_format({"bg_color": "#FFEB9C", "font_color": "#9C5700"}),
        }

    def _add_cover_sheet_xlsxwriter(self, title: str) -> None:
        """Add cover sheet to Excel workbook.

        Args:
            title: Report title
        """
        assert self.workbook is not None, "Workbook must be initialized"
        worksheet = self.workbook.add_worksheet("Cover")

        # Title
        worksheet.merge_range("B2:F2", title, self.formats["title"])

        # Report info
        worksheet.write("B4", "Generated:", self.formats["text"])
        worksheet.write("C4", datetime.now(), self.formats["date"])

        worksheet.write("B5", "Report Type:", self.formats["text"])
        worksheet.write("C5", "Financial Statements & Analysis", self.formats["text"])

        # Contents
        worksheet.write("B7", "Contents:", self.formats["section_header"])
        row = 8

        if self.config.include_balance_sheet:
            worksheet.write(f"B{row}", "• Balance Sheet", self.formats["text"])
            row += 1

        if self.config.include_income_statement:
            worksheet.write(f"B{row}", "• Income Statement", self.formats["text"])
            row += 1

        if self.config.include_cash_flow:
            worksheet.write(f"B{row}", "• Cash Flow Statement", self.formats["text"])
            row += 1

        if self.config.include_reconciliation:
            worksheet.write(f"B{row}", "• Reconciliation Report", self.formats["text"])
            row += 1

        if self.config.include_metrics_dashboard:
            worksheet.write(f"B{row}", "• Metrics Dashboard", self.formats["text"])
            row += 1

        if self.config.include_pivot_data:
            worksheet.write(f"B{row}", "• Pivot Data", self.formats["text"])
            row += 1

        # Set column widths
        worksheet.set_column("A:A", 5)
        worksheet.set_column("B:B", 20)
        worksheet.set_column("C:F", 15)

    def _write_balance_sheets_xlsxwriter(self, generator: FinancialStatementGenerator) -> None:
        """Write balance sheets to Excel workbook.

        Args:
            generator: Financial statement generator
        """
        assert self.workbook is not None, "Workbook must be initialized"
        worksheet = self.workbook.add_worksheet("Balance Sheet")

        # Title
        worksheet.merge_range("A1:F1", "BALANCE SHEET", self.formats["title"])

        # Generate balance sheets for all years
        col_offset = 0
        for year in range(min(5, generator.years_available)):  # Show up to 5 years
            df = generator.generate_balance_sheet(year)

            # Write headers
            row = 3
            worksheet.write(row, col_offset, "Item", self.formats["header"])
            worksheet.write(row, col_offset + 1, f"Year {year}", self.formats["header"])

            # Write data
            for idx, data_row in df.iterrows():
                row += 1
                item = data_row["Item"]
                value = data_row[f"Year {year}"]
                row_type = data_row.get("Type", "")

                # Format based on type
                if row_type == "total":
                    item_format = self.formats["section_header"]
                    value_format = self.formats["total"]
                elif row_type == "subtotal":
                    item_format = self.formats["text"]
                    value_format = self.formats["subtotal"]
                elif item.startswith("  "):
                    item_format = self.formats["text_indent"]
                    value_format = self.formats["currency"]
                elif item.strip() in ["ASSETS", "LIABILITIES", "EQUITY"]:
                    item_format = self.formats["section_header"]
                    value_format = None
                else:
                    item_format = self.formats["text"]
                    value_format = self.formats["currency"] if value != "" else None

                worksheet.write(row, col_offset, item, item_format)
                if value_format and value != "":
                    worksheet.write(row, col_offset + 1, value, value_format)

            col_offset += 3

        # Set column widths
        worksheet.set_column("A:A", 30)
        for col in range(1, col_offset, 3):
            worksheet.set_column(col, col, 15)
            worksheet.set_column(col + 1, col + 1, 12)

    def _write_income_statements_xlsxwriter(self, generator: FinancialStatementGenerator) -> None:
        """Write income statements to Excel workbook.

        Args:
            generator: Financial statement generator
        """
        assert self.workbook is not None, "Workbook must be initialized"
        worksheet = self.workbook.add_worksheet("Income Statement")

        # Title
        worksheet.merge_range("A1:F1", "INCOME STATEMENT", self.formats["title"])

        # Generate income statements for all years
        col_offset = 0
        for year in range(min(5, generator.years_available)):
            df = generator.generate_income_statement(year)

            # Write headers
            row = 3
            worksheet.write(row, col_offset, "Item", self.formats["header"])
            worksheet.write(row, col_offset + 1, f"Year {year}", self.formats["header"])

            # Write data
            for idx, data_row in df.iterrows():
                row += 1
                item = data_row["Item"]
                value = data_row[f"Year {year}"]
                unit = data_row.get("Unit", "")
                row_type = data_row.get("Type", "")

                # Format based on type and unit
                if row_type == "total":
                    item_format = self.formats["section_header"]
                    value_format = self.formats["total"]
                elif row_type == "subtotal":
                    item_format = self.formats["text"]
                    value_format = self.formats["subtotal"]
                elif unit == "%":
                    item_format = self.formats["text_indent"]
                    value_format = self.formats["percent"]
                elif item.startswith("  "):
                    item_format = self.formats["text_indent"]
                    value_format = self.formats["currency"]
                elif item.strip() in ["REVENUE", "OPERATING EXPENSES", "OTHER INCOME (EXPENSES)"]:
                    item_format = self.formats["section_header"]
                    value_format = None
                else:
                    item_format = self.formats["text"]
                    value_format = self.formats["currency"] if value != "" else None

                worksheet.write(row, col_offset, item, item_format)
                if value_format and value != "":
                    if unit == "%":
                        worksheet.write(row, col_offset + 1, value / 100, value_format)
                    else:
                        worksheet.write(row, col_offset + 1, value, value_format)

            col_offset += 3

        # Set column widths
        worksheet.set_column("A:A", 30)
        for col in range(1, col_offset, 3):
            worksheet.set_column(col, col, 15)
            worksheet.set_column(col + 1, col + 1, 12)

    def _write_cash_flows_xlsxwriter(self, generator: FinancialStatementGenerator) -> None:
        """Write cash flow statements to Excel workbook.

        Args:
            generator: Financial statement generator
        """
        assert self.workbook is not None, "Workbook must be initialized"
        worksheet = self.workbook.add_worksheet("Cash Flow")

        # Title
        worksheet.merge_range("A1:F1", "CASH FLOW STATEMENT", self.formats["title"])

        # Generate cash flow statements
        col_offset = 0
        for year in range(min(5, generator.years_available)):
            df = generator.generate_cash_flow_statement(year)

            # Write headers
            row = 3
            worksheet.write(row, col_offset, "Item", self.formats["header"])
            worksheet.write(row, col_offset + 1, f"Year {year}", self.formats["header"])

            # Write data
            for idx, data_row in df.iterrows():
                row += 1
                item = data_row["Item"]
                value = data_row[f"Year {year}"]
                row_type = data_row.get("Type", "")

                # Format based on type
                if row_type == "total":
                    item_format = self.formats["section_header"]
                    value_format = self.formats["total"]
                elif row_type == "subtotal":
                    item_format = self.formats["text"]
                    value_format = self.formats["subtotal"]
                elif item.startswith("    "):
                    item_format = self.formats["text_indent"]
                    value_format = self.formats["currency"]
                elif item.startswith("  "):
                    item_format = self.formats["text"]
                    value_format = self.formats["currency"]
                elif item.strip() in [
                    "OPERATING ACTIVITIES",
                    "INVESTING ACTIVITIES",
                    "FINANCING ACTIVITIES",
                ]:
                    item_format = self.formats["section_header"]
                    value_format = None
                else:
                    item_format = self.formats["text"]
                    value_format = self.formats["currency"] if value != "" else None

                worksheet.write(row, col_offset, item, item_format)
                if value_format and value != "":
                    worksheet.write(row, col_offset + 1, value, value_format)

            col_offset += 3

        # Set column widths
        worksheet.set_column("A:A", 35)
        for col in range(1, col_offset, 3):
            worksheet.set_column(col, col, 15)
            worksheet.set_column(col + 1, col + 1, 12)

    def _write_reconciliation_xlsxwriter(self, generator: FinancialStatementGenerator) -> None:
        """Write reconciliation report to Excel workbook.

        Args:
            generator: Financial statement generator
        """
        assert self.workbook is not None, "Workbook must be initialized"
        worksheet = self.workbook.add_worksheet("Reconciliation")

        # Title
        worksheet.merge_range("A1:E1", "RECONCILIATION REPORT", self.formats["title"])

        # Generate reconciliation for each year
        row = 3
        for year in range(generator.years_available):
            # Year header
            worksheet.merge_range(row, 0, row, 4, f"Year {year}", self.formats["header"])
            row += 1

            # Generate reconciliation
            df = generator.generate_reconciliation_report(year)

            # Write headers
            worksheet.write(row, 0, "Check", self.formats["subheader"])
            worksheet.write(row, 1, "Value", self.formats["subheader"])
            worksheet.write(row, 2, "Expected", self.formats["subheader"])
            worksheet.write(row, 3, "Status", self.formats["subheader"])
            row += 1

            # Write data
            for idx, data_row in df.iterrows():
                check = data_row["Check"]
                value = data_row["Value"]
                expected = data_row.get("Expected", "")
                status_type = data_row.get("Type", "")

                # Format based on status
                if status_type == "status":
                    status = value
                    if (
                        "BALANCED" in str(status)
                        or "MATCHED" in str(status)
                        or "VALID" in str(status)
                        or "SOLVENT" in str(status)
                    ):
                        status_format = self.formats["good"]
                    elif (
                        "IMBALANCED" in str(status)
                        or "MISMATCHED" in str(status)
                        or "INVALID" in str(status)
                        or "INSOLVENT" in str(status)
                    ):
                        status_format = self.formats["bad"]
                    else:
                        status_format = self.formats["neutral"]

                    worksheet.write(row, 0, check, self.formats["text"])
                    worksheet.write(row, 3, status, status_format)
                else:
                    worksheet.write(row, 0, check, self.formats["text"])
                    if value != "":
                        worksheet.write(row, 1, value, self.formats["number"])
                    if expected != "":
                        worksheet.write(row, 2, expected, self.formats["number"])

                row += 1

            row += 2  # Space between years

        # Set column widths
        worksheet.set_column("A:A", 30)
        worksheet.set_column("B:C", 15)
        worksheet.set_column("D:D", 12)

    def _write_metrics_dashboard_xlsxwriter(  # pylint: disable=too-many-branches
        self, generator: FinancialStatementGenerator
    ) -> None:
        """Write metrics dashboard to Excel workbook.

        Args:
            generator: Financial statement generator

        Note: Multiple branches are necessary for proper Excel formatting
        of different data types (currency, percentages, numbers, etc.)
        """
        assert self.workbook is not None, "Workbook must be initialized"
        worksheet = self.workbook.add_worksheet("Metrics Dashboard")

        # Title
        worksheet.merge_range("A1:H1", "KEY METRICS DASHBOARD", self.formats["title"])

        # Prepare metrics data
        metrics_data = []
        for year in range(generator.years_available):
            metrics = generator.metrics_history[year]
            metrics_data.append(
                {
                    "Year": year,
                    "Revenue": metrics.get("revenue", 0),
                    "Operating Income": metrics.get("operating_income", 0),
                    "Net Income": metrics.get("net_income", 0),
                    "Assets": metrics.get("assets", 0),
                    "Equity": metrics.get("equity", 0),
                    "ROE %": metrics.get("roe", 0) * 100,
                    "ROA %": metrics.get("roa", 0) * 100,
                    "Operating Margin %": metrics.get("operating_margin", 0) * 100,
                    "Asset Turnover": metrics.get("asset_turnover", 0),
                    "Collateral": metrics.get("collateral", 0),
                    "Claim Liabilities": metrics.get("claim_liabilities", 0),
                    "Solvent": "Yes" if metrics.get("is_solvent", True) else "No",
                }
            )

        # Create DataFrame
        df = pd.DataFrame(metrics_data)

        # Write headers
        row = 3
        for col, header in enumerate(df.columns):
            worksheet.write(row, col, header, self.formats["header"])

        # Write data
        for idx, data_row in df.iterrows():
            row += 1
            for col, (key, value) in enumerate(data_row.items()):
                if key == "Year":
                    worksheet.write(row, col, value, self.formats["number"])
                elif key in [
                    "Revenue",
                    "Operating Income",
                    "Net Income",
                    "Assets",
                    "Equity",
                    "Collateral",
                    "Claim Liabilities",
                ]:
                    worksheet.write(row, col, value, self.formats["currency"])
                elif "%" in str(key):
                    worksheet.write(row, col, value / 100, self.formats["percent"])
                elif key == "Solvent":
                    format_to_use = self.formats["good"] if value == "Yes" else self.formats["bad"]
                    worksheet.write(row, col, value, format_to_use)
                else:
                    worksheet.write(row, col, value, self.formats["number"])

        # Add summary statistics
        row += 3
        worksheet.merge_range(row, 0, row, 2, "SUMMARY STATISTICS", self.formats["header"])
        row += 1

        # Calculate and write summary stats
        summary_stats = [
            ("Average Revenue", df["Revenue"].mean()),
            (
                "Revenue CAGR",
                ((df["Revenue"].iloc[-1] / df["Revenue"].iloc[0]) ** (1 / len(df)) - 1) * 100
                if len(df) > 1
                else 0,
            ),
            ("Average ROE %", df["ROE %"].mean()),
            ("Average ROA %", df["ROA %"].mean()),
            ("Max Collateral", df["Collateral"].max()),
            ("Max Claim Liabilities", df["Claim Liabilities"].max()),
        ]

        for stat_name, stat_value in summary_stats:
            worksheet.write(row, 0, stat_name, self.formats["text"])
            if "%" in stat_name or "CAGR" in stat_name:
                worksheet.write(row, 1, stat_value / 100, self.formats["percent"])
            elif "Revenue" in stat_name or "Collateral" in stat_name or "Claim" in stat_name:
                worksheet.write(row, 1, stat_value, self.formats["currency"])
            else:
                worksheet.write(row, 1, stat_value, self.formats["number"])
            row += 1

        # Set column widths
        worksheet.set_column("A:A", 15)
        worksheet.set_column("B:L", 18)
        worksheet.set_column("M:M", 10)

        # Add conditional formatting for trends
        if len(df) > 1:
            # Highlight positive/negative growth
            worksheet.conditional_format(
                f"G5:I{4 + len(df)}",
                {"type": "cell", "criteria": ">", "value": 0, "format": self.formats["good"]},
            )

    def _write_pivot_data_xlsxwriter(self, generator: FinancialStatementGenerator) -> None:
        """Write pivot-ready data to Excel workbook.

        Args:
            generator: Financial statement generator
        """
        assert self.workbook is not None, "Workbook must be initialized"
        worksheet = self.workbook.add_worksheet("Pivot Data")

        # Prepare normalized data for pivot tables
        pivot_data = []

        for year in range(generator.years_available):
            metrics = generator.metrics_history[year]

            # Financial metrics
            for metric_name, metric_value in metrics.items():
                if isinstance(metric_value, (int, float)):
                    pivot_data.append(
                        {
                            "Year": year,
                            "Category": self._categorize_metric(metric_name),
                            "Metric": metric_name.replace("_", " ").title(),
                            "Value": metric_value,
                        }
                    )

        # Create DataFrame
        df = pd.DataFrame(pivot_data)

        # Write to worksheet
        row = 0

        # Write headers
        for col, header in enumerate(df.columns):
            worksheet.write(row, col, header, self.formats["header"])

        # Write data
        for idx, data_row in df.iterrows():
            row += 1
            for col, value in enumerate(data_row):
                if col == 0:  # Year
                    worksheet.write(row, col, value, self.formats["number"])
                elif col == 3:  # Value
                    worksheet.write(row, col, value, self.formats["number"])
                else:  # Category, Metric
                    worksheet.write(row, col, value, self.formats["text"])

        # Create pivot table if possible
        if XLSXWRITER_AVAILABLE and len(df) > 0:
            # Add a pivot table worksheet
            assert self.workbook is not None
            pivot_worksheet = self.workbook.add_worksheet("Pivot Analysis")

            # Create pivot table
            pivot_worksheet.add_table(
                f"A3:D{3 + len(df)}",
                {
                    "data": df.values.tolist(),
                    "columns": [{"header": col} for col in df.columns],
                    "style": "Table Style Light 1",
                },
            )

            # Add instructions
            pivot_worksheet.merge_range(
                "A1:D1",
                "Use this data to create pivot tables for custom analysis",
                self.formats["title"],
            )

        # Set column widths
        worksheet.set_column("A:A", 10)
        worksheet.set_column("B:B", 20)
        worksheet.set_column("C:C", 25)
        worksheet.set_column("D:D", 15)

    def _categorize_metric(self, metric_name: str) -> str:
        """Categorize a metric for pivot table grouping.

        Args:
            metric_name: Name of the metric

        Returns:
            Category name
        """
        if any(x in metric_name for x in ["revenue", "income", "profit"]):
            return "Income"
        if any(x in metric_name for x in ["asset", "equity", "collateral"]):
            return "Balance Sheet"
        if any(x in metric_name for x in ["roe", "roa", "margin", "turnover"]):
            return "Ratios"
        if any(x in metric_name for x in ["claim", "liability"]):
            return "Liabilities"
        return "Other"

    def _generate_with_openpyxl(
        self, generator: FinancialStatementGenerator, output_path: Path, title: Optional[str] = None
    ) -> None:
        """Generate report using openpyxl engine.

        Args:
            generator: Financial statement generator
            output_path: Output file path
            title: Optional report title
        """
        import openpyxl
        from openpyxl import Workbook

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Add sheets with data
        if self.config.include_balance_sheet:
            self._add_balance_sheet_openpyxl(wb, generator)

        if self.config.include_income_statement:
            self._add_income_statement_openpyxl(wb, generator)

        if self.config.include_cash_flow:
            self._add_cash_flow_openpyxl(wb, generator)

        if self.config.include_reconciliation:
            self._add_reconciliation_openpyxl(wb, generator)

        if self.config.include_metrics_dashboard:
            self._add_metrics_dashboard_openpyxl(wb, generator)

        # Save workbook
        wb.save(str(output_path))

    def _add_balance_sheet_openpyxl(self, wb: Any, generator: FinancialStatementGenerator) -> None:
        """Add balance sheet to openpyxl workbook.

        Args:
            wb: Openpyxl workbook
            generator: Financial statement generator
        """
        ws = wb.create_sheet("Balance Sheet")

        # Generate balance sheet for latest year
        year = generator.years_available - 1
        df = generator.generate_balance_sheet(year)

        # Write data directly as a proper table that pandas can read
        # Write headers in row 1
        ws["A1"] = "Item"
        ws["B1"] = f"Year {year}"
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)

        # Write data starting from row 2
        row = 2
        for idx, data_row in df.iterrows():
            ws[f"A{row}"] = data_row["Item"]
            value = data_row[f"Year {year}"]
            if value != "" and value is not None:
                ws[f"B{row}"] = value
            row += 1

        # Format columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

    def _add_income_statement_openpyxl(
        self, wb: Any, generator: FinancialStatementGenerator
    ) -> None:
        """Add income statement to openpyxl workbook.

        Args:
            wb: Openpyxl workbook
            generator: Financial statement generator
        """
        ws = wb.create_sheet("Income Statement")

        # Generate income statement for latest year
        year = generator.years_available - 1
        df = generator.generate_income_statement(year)

        # Write title
        ws["A1"] = "INCOME STATEMENT"
        ws["A1"].font = Font(bold=True, size=14)

        # Write headers
        ws["A3"] = "Item"
        ws["B3"] = f"Year {year}"

        # Write data
        row = 4
        for idx, data_row in df.iterrows():
            ws[f"A{row}"] = data_row["Item"]
            value = data_row[f"Year {year}"]
            if value != "" and value is not None:
                ws[f"B{row}"] = value
            row += 1

        # Format columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

    def _add_cash_flow_openpyxl(self, wb: Any, generator: FinancialStatementGenerator) -> None:
        """Add cash flow statement to openpyxl workbook.

        Args:
            wb: Openpyxl workbook
            generator: Financial statement generator
        """
        ws = wb.create_sheet("Cash Flow")

        # Generate cash flow for latest year
        year = generator.years_available - 1
        df = generator.generate_cash_flow_statement(year)

        # Write title
        ws["A1"] = "CASH FLOW STATEMENT"
        ws["A1"].font = Font(bold=True, size=14)

        # Write headers
        ws["A3"] = "Item"
        ws["B3"] = f"Year {year}"

        # Write data
        row = 4
        for idx, data_row in df.iterrows():
            ws[f"A{row}"] = data_row["Item"]
            value = data_row[f"Year {year}"]
            if value != "" and value is not None:
                ws[f"B{row}"] = value
            row += 1

        # Format columns
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15

    def _add_reconciliation_openpyxl(self, wb: Any, generator: FinancialStatementGenerator) -> None:
        """Add reconciliation report to openpyxl workbook.

        Args:
            wb: Openpyxl workbook
            generator: Financial statement generator
        """
        ws = wb.create_sheet("Reconciliation")

        # Generate reconciliation for latest year
        year = generator.years_available - 1
        df = generator.generate_reconciliation_report(year)

        # Write title
        ws["A1"] = "RECONCILIATION REPORT"
        ws["A1"].font = Font(bold=True, size=14)

        # Write headers
        ws["A3"] = "Check"
        ws["B3"] = "Value"
        ws["C3"] = "Expected"
        ws["D3"] = "Status"

        # Write data
        row = 4
        for idx, data_row in df.iterrows():
            ws[f"A{row}"] = data_row["Check"]
            if data_row["Value"] != "":
                ws[f"B{row}"] = data_row["Value"]
            if data_row.get("Expected", "") != "":
                ws[f"C{row}"] = data_row["Expected"]
            if data_row.get("Type") == "status":
                ws[f"D{row}"] = data_row["Value"]
            row += 1

        # Format columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12

    def _add_metrics_dashboard_openpyxl(
        self, wb: Any, generator: FinancialStatementGenerator
    ) -> None:
        """Add metrics dashboard to openpyxl workbook.

        Args:
            wb: Openpyxl workbook
            generator: Financial statement generator
        """
        ws = wb.create_sheet("Metrics Dashboard")

        # Prepare metrics data
        headers = [
            "Year",
            "Revenue",
            "Operating Income",
            "Net Income",
            "Assets",
            "Equity",
            "ROE %",
            "ROA %",
            "Operating Margin %",
        ]

        # Write headers in row 1 for pandas compatibility
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)

        # Write data starting from row 2
        row = 2
        for year in range(generator.years_available):
            metrics = generator.metrics_history[year]
            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=2, value=metrics.get("revenue", 0))
            ws.cell(row=row, column=3, value=metrics.get("operating_income", 0))
            ws.cell(row=row, column=4, value=metrics.get("net_income", 0))
            ws.cell(row=row, column=5, value=metrics.get("assets", 0))
            ws.cell(row=row, column=6, value=metrics.get("equity", 0))
            ws.cell(row=row, column=7, value=metrics.get("roe", 0) * 100)
            ws.cell(row=row, column=8, value=metrics.get("roa", 0) * 100)
            ws.cell(row=row, column=9, value=metrics.get("operating_margin", 0) * 100)
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _generate_with_pandas(
        self, generator: FinancialStatementGenerator, output_path: Path
    ) -> None:
        """Generate report using pandas default Excel writer.

        Args:
            generator: Financial statement generator
            output_path: Output file path
        """
        with pd.ExcelWriter(
            str(output_path), engine="openpyxl" if OPENPYXL_AVAILABLE else None
        ) as writer:
            # Write balance sheet
            if self.config.include_balance_sheet:
                year = generator.years_available - 1
                df = generator.generate_balance_sheet(year)
                df.to_excel(writer, sheet_name="Balance Sheet", index=False)

            # Write income statement
            if self.config.include_income_statement:
                year = generator.years_available - 1
                df = generator.generate_income_statement(year)
                df.to_excel(writer, sheet_name="Income Statement", index=False)

            # Write cash flow
            if self.config.include_cash_flow:
                year = generator.years_available - 1
                df = generator.generate_cash_flow_statement(year)
                df.to_excel(writer, sheet_name="Cash Flow", index=False)

            # Write reconciliation
            if self.config.include_reconciliation:
                year = generator.years_available - 1
                df = generator.generate_reconciliation_report(year)
                df.to_excel(writer, sheet_name="Reconciliation", index=False)

            # Write metrics dashboard
            if self.config.include_metrics_dashboard:
                metrics_data = []
                for year in range(generator.years_available):
                    metrics = generator.metrics_history[year]
                    metrics_data.append(
                        {
                            "Year": year,
                            "Revenue": metrics.get("revenue", 0),
                            "Operating Income": metrics.get("operating_income", 0),
                            "Net Income": metrics.get("net_income", 0),
                            "Assets": metrics.get("assets", 0),
                            "Equity": metrics.get("equity", 0),
                            "ROE %": metrics.get("roe", 0) * 100,
                            "ROA %": metrics.get("roa", 0) * 100,
                            "Operating Margin %": metrics.get("operating_margin", 0) * 100,
                        }
                    )
                df = pd.DataFrame(metrics_data)
                df.to_excel(writer, sheet_name="Metrics Dashboard", index=False)

    def generate_monte_carlo_report(
        self,
        results: Any,  # TODO: Replace Any with MonteCarloResults when implemented  # pylint: disable=fixme
        output_file: str,
        title: Optional[str] = None,
    ) -> Path:
        """Generate aggregated report from Monte Carlo simulations.

        Creates Excel report with statistical summaries across multiple
        simulation trajectories.

        Args:
            results: Monte Carlo simulation results
            output_file: Name of output Excel file
            title: Optional report title

        Returns:
            Path to generated Excel file
        """
        output_path = self.config.output_path / output_file

        # Create a placeholder Excel file for now
        # Full implementation will come when MonteCarloResults is available
        with pd.ExcelWriter(
            str(output_path), engine="openpyxl" if OPENPYXL_AVAILABLE else None
        ) as writer:
            # Create a placeholder summary sheet
            placeholder_df = pd.DataFrame(
                {
                    "Note": ["Monte Carlo Report Placeholder"],
                    "Status": ["To be implemented with MonteCarloResults"],
                }
            )
            placeholder_df.to_excel(writer, sheet_name="Summary", index=False)

        return output_path
